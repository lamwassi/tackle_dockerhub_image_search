import json
from openpyxl import load_workbook
import re

def cleanStrValue(value):

    """
    arguments:   
        value : str 
    Return:             
        str : formated str
    """    
    if value:
        value = str(value).strip()
        value = value.replace(u'\u00a0', ' ')
        value = re.sub(r'[^\x00-\x7F]+', ' ', ' ' + str(value) + ' ').strip()
    else:
        value = ''
        
    return value

def load_entity(inputexcel , sheetname = "docker_images"):

    if not inputexcel:
        return None

    sheet = load_workbook(filename = inputexcel)
    sheet_data = sheet[sheetname]
    max_row = sheet_data.max_row
    
    entity = {}
    for xint in range(2, max_row+1):
        row_num = str(xint) 
        entity_name = cleanStrValue(value = sheet_data["B" + row_num].value)
        entity_id = xint-1
        entity[str(entity_id)] = entity_name
    return entity


def load_images(inputexcel , sheetname = "docker_images"):

    """
    Args:
        inputexcel:
    Returns:
        dict: dictionary of dockerimage name entity as key and url as value
    """

    if not inputexcel:
        return None


    sheet = load_workbook(filename = inputexcel)
    sheet_data = sheet[sheetname]
    max_row = sheet_data.max_row
    entities = []
    empty_dic = {}
    for xint in range(2, max_row+1):
        row_num = str(xint) 
        new_dic = empty_dic.copy()
        new_dic["entity_name"] = cleanStrValue(value = sheet_data[ 'B' + row_num].value)
        new_dic["app"] = cleanStrValue(value = sheet_data["F" + row_num].value)
        new_dic["app_server"] = cleanStrValue(value = sheet_data["G"+ row_num].value)
       
        entities.append(new_dic)
       
    return entities



def filter_entity(entity_names1, entity_names2):
    """
    """

    ent_name1=  [name1.lower() for _ , name1 in entity_names1.items()]
    ent_name2 = [name2.lower() for _, name2 in entity_names2.items()]
    ent_name = list(set(ent_name1 + ent_name2))
    
    entities = {}
    for i , name in enumerate(ent_name):
        entities[str(i)] = name
    
    return entities